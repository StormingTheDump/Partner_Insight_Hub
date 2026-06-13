from fastapi import FastAPI, File, Header, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from database import init_tables, get_connection
from auth import login, verify_token, hash_password
import os
from typing import Optional

app = FastAPI(title="Partner Insight Hub API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "http://10.6.20.87:5173",
        "http://198.19.107.12:5173",
        "https://huangxiaozhen.github.io",
    ],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def startup():
    init_tables()


class LoginRequest(BaseModel):
    email: str
    password: str


class CreateUserRequest(BaseModel):
    email: str
    password: str
    channel_name: str
    contact_name: str


class ChatMessage(BaseModel):
    role: str
    content: str


class ChatRequest(BaseModel):
    messages: list[ChatMessage]


ERROR_STATUS = {"not_found": 404, "wrong_password": 401, "disabled": 403}
ERROR_MSG = {
    "not_found": "账号不存在，请检查邮箱地址",
    "wrong_password": "密码错误，请重新输入",
    "disabled": "该账号已被禁用，请联系客户经理",
}

DIDA_API_SYSTEM_PROMPT = """You are a helpful assistant specialized in Dida Travel API (DidaAPI). Answer questions clearly and concisely in the same language the user uses (Chinese or English).

## Dida API Overview
- API Version: v2.0
- Base URL (Test): https://apiint.didatravel.com
- Base URL (Production): https://api.didatravel.com
- Contact: techsupport@dida.com
- All requests use HTTP POST
- Supports JSON and XML (JSON recommended)
- Default QPS limit: 30 (configurable)
- All responses are gzip encoded

## Authentication
Every request must include a Header with ClientID and LicenseKey:
```json
{
  "Header": {
    "ClientID": "YourClientID",
    "LicenseKey": "YourLicenseKey"
  }
}
```

## Go Live Process
1. Sign NDA
2. Access development environment (test credentials provided)
3. Complete Post Sale Questionnaire
4. Pre-development communication with Dida integration team
5. Test credentials and pre-production test (send test template to Dida)
6. Live credentials and test booking
7. Go live

## Booking API Flow
The standard booking flow is:
1. **Price Search** → Get available rates
2. **Price Confirm** → Validate price, get booking reference
3. **Booking Confirm** → Create booking using reference from step 2
4. **Booking Search** → Check booking status
5. **Booking Cancel** (if needed) → Pre-cancel → Cancel Confirm

### 3.1 Price Search
- POST `https://apiint.didatravel.com/api/hotel/HotelPriceSearch?$format=json`
- Three types: Lowest Price Search, Cache Rate Search, Realtime Rate Search
- Lowest Price Search and Cache Rate Search return prices for 2 persons by default
- Cache is on a daily basis; for 30-day cache, invoke 30 times
- For multiple hotels, use Cache Rate Search
- PriceSearch returns price for ONE room; for multiple rooms multiply accordingly
- Supports single requests with multiple rooms of the SAME room type (same RatePlanID)
- Does NOT support multiple rooms of different room types in one request

### 3.2 Price Confirm
- POST `https://apiint.didatravel.com/api/hotel/HotelPriceConfirm?$format=json`
- Validates price and returns a booking reference (BookingReference/ReferenceNo)
- The reference is required for Booking Confirm
- Returns total price for all rooms

### 3.3 Booking Confirm
- POST `https://apiint.didatravel.com/api/booking/HotelBookingConfirm?$format=json`
- Creates a new booking using the reference from Price Confirm
- Parameters must match Price Confirm (check-in/out dates, rooms, occupancy)
- ConfirmationCode = Hotel Confirmation Number (HCN) — not always returned in real-time
- Name field supports: English, Chinese, Japanese, Korean, Portuguese, and more

### 3.4 Booking Search
- POST `https://apiint.didatravel.com/api/booking/HotelBookingSearch?$format=json`
- Search by BookingID (recommended) or other params
- **Final booking statuses only**: Status 2 (Confirmed), 3 (Canceled), 4 (Failed)
- Do NOT interpret empty responses or other statuses as final

**Booking Status Codes:**
| Status | Name | Description |
|--------|------|-------------|
| 0 | PreBook | - |
| 1 | Booked | PreBook + awaiting payment |
| 2 | Confirmed | **Final** — booking confirmed |
| 3 | Canceled | **Final** — booking cancelled |
| 4 | Failed | **Final** — booking failed |
| 5 | Pending | Transition phase (within ~3 minutes) |
| 6 | OnRequest | May take up to 120 minutes to finalize |

### 3.5 Booking Cancel
Two-step process:
1. **Pre-cancel**: POST `.../HotelBookingCancel?$format=json` — returns CancelConfirmID (valid 10 minutes)
2. **Cancel Confirm**: POST `.../HotelBookingCancelConfirm?$format=json` — confirms cancellation using CancelConfirmID
- If CancelConfirmID expires, repeat Pre-cancel to get a new one

## Content API
- Hotel list, hotel details, room types, images, facilities, bed types, etc.
- Content API v2 available with enhanced fields
- Used to build local hotel content cache

## Error Codes

### Pricing Errors (2xxx)
| Code | Description |
|------|-------------|
| 2000 | Unexpected Error |
| 2001 | No Hotel Or Destination |
| 2002 | No Hotel Found |
| 2003 | Incorrect Date |
| 2004 | Incorrect RoomCount |
| 2005 | No Available Room |
| 2006 | No RatePlan Found |
| 2009 | Check Availability Timeout |
| 2017 | Client Auth Failed |
| 2018 | Currency Not Supported |
| 2022 | Limited Call (Exceed QPS limitation) |
| 2029 | Hotel Stop Sell |

### Booking Errors (3xxx)
| Code | Description |
|------|-------------|
| 3000 | Unexpected Error |
| 3001 | Incorrect Booking Information |
| 3002 | Incorrect ReferenceNo |
| 3003 | Incorrect BookingID |
| 3004 | Incorrect CancelConfirmID |
| 3006 | Booking Expired |
| 3007 | Cancel Expired |
| 3014 | Not Enough Credit |
| 3015 | Availability Or Price Invalid |
| 3016 | Failed To Confirm Booking |
| 3019 | Duplicate ClientReference |

## Best Practices
- Use Cache Rate Search for multiple hotel queries
- Always use Booking Search with BookingID for status checks
- Call Booking Search 3 days before check-in to get HCN if not returned in real-time
- Handle Pending (5) and OnRequest (6) statuses with polling
- Implement timeout settings according to Dida's recommendation
- White-list your server IP with Dida
- Request higher QPS if default 30 is not enough

Answer questions based on this documentation. If asked something outside this scope, say so honestly.
"""


# ── Agoda Metrics API ─────────────────────────────────────────────────────────

@app.get("/api/metrics/funnel")
def metrics_funnel(
    start_date: Optional[str] = Query(default=None),
    end_date:   Optional[str] = Query(default=None),
):
    """
    5层转化漏斗：查价数 → 有价数 → 验价数 → 准确验价数 → 下单数
    """
    conn = get_connection()

    date_cond_s = ""
    date_cond_c = ""
    date_cond_m = ""
    date_params: list = []
    if start_date and end_date:
        date_cond_s = "WHERE date BETWEEN ? AND ?"
        date_cond_c = "WHERE date BETWEEN ? AND ?"
        date_cond_m = "WHERE date BETWEEN ? AND ?"
        date_params = [start_date, end_date]

    overall = conn.execute(f"""
        SELECT
          (SELECT SUM(search_requests)  FROM agoda_price_search  {date_cond_s}) AS searches,
          (SELECT SUM(result_count)     FROM agoda_price_search  {date_cond_s}) AS results,
          (SELECT SUM(confirm_requests) FROM agoda_price_confirm {date_cond_c}) AS confirms,
          (SELECT SUM(accurate_count)   FROM agoda_price_confirm {date_cond_c}) AS accurates,
          (SELECT SUM(bookings)         FROM agoda_daily_metrics {date_cond_m}) AS bookings,
          (SELECT ROUND(AVG(accurate_rate),1)  FROM agoda_price_confirm {date_cond_c}) AS confirm_accurate_rate,
          (SELECT ROUND(AVG(avg_response_ms))  FROM agoda_price_search  {date_cond_s}) AS avg_response_ms
    """, date_params * 7 if date_params else []).fetchone()

    join_where = f"WHERE s.date BETWEEN ? AND ?" if (start_date and end_date) else ""
    by_client = conn.execute(f"""
        SELECT
            s.client_id,
            SUM(s.search_requests)                                                  AS searches,
            SUM(s.result_count)                                                     AS results,
            SUM(c.confirm_requests)                                                 AS confirms,
            SUM(c.accurate_count)                                                   AS accurates,
            SUM(m.bookings)                                                         AS bookings,
            ROUND(SUM(s.result_count)*100.0/NULLIF(SUM(s.search_requests),0),1)    AS result_rate,
            ROUND(SUM(c.confirm_requests)*100.0/NULLIF(SUM(s.search_requests),0),1) AS search_to_confirm_rate,
            ROUND(SUM(c.accurate_count)*100.0/NULLIF(SUM(c.confirm_requests),0),1) AS accurate_rate,
            ROUND(SUM(m.bookings)*100.0/NULLIF(SUM(c.confirm_requests),0),1)       AS confirm_to_book_rate,
            ROUND(AVG(s.avg_response_ms))                                           AS avg_response_ms
        FROM agoda_price_search  s
        JOIN agoda_price_confirm c ON s.date = c.date AND s.client_id = c.client_id
        JOIN agoda_daily_metrics m ON s.date = m.date AND s.client_id = m.client_id
        {join_where}
        GROUP BY s.client_id
        ORDER BY searches DESC
    """, ([start_date, end_date] if (start_date and end_date) else [])).fetchall()

    conn.close()

    searches = overall["searches"]
    results  = overall["results"]
    confirms = overall["confirms"]
    accurates= overall["accurates"]
    bookings = overall["bookings"]

    return {
        "overall": {
            "searches":          searches,
            "results":           results,
            "confirms":          confirms,
            "accurates":         accurates,
            "bookings":          bookings,
            "result_rate":       round(results   / searches * 100, 1),
            "search_to_confirm": round(confirms  / searches * 100, 1),
            "accurate_rate":     round(accurates / confirms * 100, 1),
            "confirm_to_book":   round(bookings  / confirms * 100, 1),
            "avg_response_ms":   overall["avg_response_ms"],
        },
        "by_client": [dict(r) for r in by_client],
    }


@app.get("/api/metrics/overview")
def metrics_overview(
    client_id:  str = Query(default=None),
    start_date: str = Query(default=None),
    end_date:   str = Query(default=None),
):
    """
    汇总卡片 + 30天日时序（供概览页折线图/sparkline使用）
    """
    conn = get_connection()
    conditions = []
    params: list = []
    if client_id:
        conditions.append("m.client_id = ?")
        params.append(client_id)
    if start_date and end_date:
        conditions.append("m.date BETWEEN ? AND ?")
        params.extend([start_date, end_date])
    where_m = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    # 汇总卡片
    summary = conn.execute(f"""
        SELECT
            SUM(m.bookings)                                              AS total_bookings,
            ROUND(SUM(m.ttv), 2)                                         AS total_ttv,
            ROUND(SUM(m.ttv) / NULLIF(SUM(m.bookings), 0), 2)           AS avg_order_value,
            SUM(m.room_nights)                                           AS total_room_nights,
            ROUND(AVG(m.pre_error_rate), 2)                              AS avg_pre_error_rate,
            ROUND(AVG(m.book_error_rate), 2)                             AS avg_book_error_rate,
            ROUND(AVG(s.avg_response_ms))                                AS avg_response_ms
        FROM agoda_daily_metrics m
        LEFT JOIN agoda_price_search s ON m.date = s.date AND m.client_id = s.client_id
        {where_m}
    """, params).fetchone()

    # 日时序
    daily = conn.execute(f"""
        SELECT
            m.date,
            SUM(m.bookings)                                              AS bookings,
            ROUND(SUM(m.ttv) / 1000.0, 1)                               AS ttv_k,
            ROUND(SUM(m.ttv) / NULLIF(SUM(m.bookings), 0), 2)           AS avg_order_value,
            SUM(m.room_nights)                                           AS room_nights,
            ROUND(AVG(m.pre_error_rate), 2)                              AS pre_error_rate,
            ROUND(AVG(m.book_error_rate), 2)                             AS book_error_rate,
            ROUND(AVG(s.avg_response_ms))                                AS avg_response_ms
        FROM agoda_daily_metrics m
        LEFT JOIN agoda_price_search s ON m.date = s.date AND m.client_id = s.client_id
        {where_m}
        GROUP BY m.date
        ORDER BY m.date
    """, params).fetchall()

    conn.close()

    dates = [r["date"] for r in daily]
    return {
        "summary": dict(summary),
        "daily": {
            "labels":          dates,
            "ttv":             [r["ttv_k"]          for r in daily],
            "bookings":        [r["bookings"]        for r in daily],
            "avg_order_value": [r["avg_order_value"] for r in daily],
            "room_nights":     [r["room_nights"]     for r in daily],
            "pre_error_rate":  [r["pre_error_rate"]  for r in daily],
            "book_error_rate": [r["book_error_rate"] for r in daily],
            "avg_response_ms": [r["avg_response_ms"] for r in daily],
        },
    }


@app.get("/api/metrics/dimensions")
def metrics_dimensions(
    client_id:  str = Query(default=None),
    start_date: str = Query(default=None),
    end_date:   str = Query(default=None),
):
    """
    订单维度细分：LT（提前预订）/ Chain（酒店类型）/ Country（目的地国家）
    """
    conn = get_connection()
    # dimension tables have client_id but no date column — filter by client_id only
    dim_conds: list = []
    dim_params: list = []
    if client_id:
        dim_conds.append("client_id = ?")
        dim_params.append(client_id)
    where = ("WHERE " + " AND ".join(dim_conds)) if dim_conds else ""

    lt = conn.execute(f"""
        SELECT lt_bucket,
               SUM(bookings)    AS bookings,
               ROUND(SUM(ttv),2) AS ttv,
               SUM(room_nights) AS room_nights
        FROM agoda_orders_by_lt {where}
        GROUP BY lt_bucket
        ORDER BY CASE lt_bucket
            WHEN '0-3天'   THEN 1 WHEN '4-7天'   THEN 2
            WHEN '8-14天'  THEN 3 WHEN '15-30天'  THEN 4
            WHEN '31+天'   THEN 5 ELSE 9 END
    """, dim_params).fetchall()

    chain = conn.execute(f"""
        SELECT chain_type,
               SUM(bookings)    AS bookings,
               ROUND(SUM(ttv),2) AS ttv,
               SUM(room_nights) AS room_nights
        FROM agoda_orders_by_chain {where}
        GROUP BY chain_type
        ORDER BY bookings DESC
    """, dim_params).fetchall()

    country = conn.execute(f"""
        SELECT country,
               SUM(bookings)    AS bookings,
               ROUND(SUM(ttv),2) AS ttv,
               SUM(room_nights) AS room_nights
        FROM agoda_orders_by_country {where}
        GROUP BY country
        ORDER BY bookings DESC
    """, dim_params).fetchall()

    star = conn.execute(f"""
        SELECT star_rating,
               SUM(bookings)     AS bookings,
               ROUND(SUM(ttv),2) AS ttv,
               SUM(room_nights)  AS room_nights
        FROM agoda_orders_by_star {where}
        GROUP BY star_rating
        ORDER BY CASE star_rating
            WHEN '0星' THEN 0 WHEN '1星' THEN 1 WHEN '2星' THEN 2
            WHEN '3星' THEN 3 WHEN '4星' THEN 4 WHEN '5星' THEN 5
            ELSE 9 END
    """, dim_params).fetchall()

    conn.close()

    def with_pct(rows, key="bookings"):
        total = sum(r[key] for r in rows) or 1
        return [{**dict(r), "pct": round(r[key] / total * 100, 1)} for r in rows]

    return {
        "lt":      with_pct(lt),
        "chain":   with_pct(chain),
        "country": with_pct(country),
        "star":    with_pct(star),
    }



@app.get("/api/metrics/performance")
def metrics_performance(
    start_date: Optional[str] = Query(default=None),
    end_date:   Optional[str] = Query(default=None),
):
    """
    各 Client ID 的聚合业绩（业绩表现页表格 + 堆叠柱状图）
    """
    conn = get_connection()
    date_cond = ""
    date_params: list = []
    if start_date and end_date:
        date_cond = "WHERE date BETWEEN ? AND ?"
        date_params = [start_date, end_date]

    # 每个 Client ID 汇总行
    rows = conn.execute(f"""
        SELECT
            client_id,
            SUM(wins)                                                    AS wins,
            SUM(opportunities)                                           AS opportunities,
            ROUND(SUM(wins) * 100.0 / NULLIF(SUM(opportunities), 0), 2) AS win_rate_pct,
            SUM(bookings)                                                AS bookings,
            ROUND(SUM(ttv), 2)                                           AS ttv,
            ROUND(SUM(ttv) / NULLIF(SUM(bookings), 0), 2)               AS avg_order_value,
            SUM(room_nights)                                             AS room_nights
        FROM agoda_daily_metrics
        {date_cond}
        GROUP BY client_id
        ORDER BY ttv DESC
    """, date_params).fetchall()

    # 按日 × Client ID（供堆叠柱状图）
    daily_by_client = conn.execute(f"""
        SELECT date, client_id, bookings
        FROM agoda_daily_metrics
        {date_cond}
        ORDER BY date, client_id
    """, date_params).fetchall()

    conn.close()

    # 整理堆叠图数据：{client_id: [day0_bookings, day1_bookings, ...]}
    from collections import defaultdict
    stacked: dict = defaultdict(list)
    dates_set: list = []
    for r in daily_by_client:
        if r["date"] not in dates_set:
            dates_set.append(r["date"])
    for client_row in rows:
        cid = client_row["client_id"]
        day_map = {r["date"]: r["bookings"] for r in daily_by_client if r["client_id"] == cid}
        stacked[cid] = [day_map.get(d, 0) for d in dates_set]

    return {
        "rows": [
            {
                "client_id":      r["client_id"],
                "wins":           r["wins"],
                "opportunities":  r["opportunities"],
                "win_rate":       f'{r["win_rate_pct"]}%',
                "bookings":       r["bookings"],
                "ttv":            r["ttv"],
                "avg_order_value":r["avg_order_value"],
                "room_nights":    r["room_nights"],
            }
            for r in rows
        ],
        "stacked": {
            "labels": dates_set,
            "series": dict(stacked),
        },
    }


# ── Auth & Chat ────────────────────────────────────────────────────────────────

@app.post("/api/auth/login")
def api_login(body: LoginRequest):
    result = login(body.email, body.password)
    if not result["success"]:
        err = result["error"]
        raise HTTPException(status_code=ERROR_STATUS[err], detail=ERROR_MSG[err])
    return result


_CHANNELS = ["Agoda", "AgodaEBK", "AgodaUK", "Lvzan", "DidaOpaq", "Barli2b"]


@app.get("/api/integration/api-metrics")
def api_integration_metrics(
    start_date: Optional[str] = Query(default=None),
    end_date:   Optional[str] = Query(default=None),
    client_id:  Optional[str] = Query(default=None),
):
    from database import get_connection
    from datetime import datetime
    from collections import defaultdict

    conn = get_connection()
    conditions = []
    params: list = []
    if start_date and end_date:
        conditions.append("date BETWEEN ? AND ?")
        params.extend([start_date, end_date])
    if client_id:
        conditions.append("channel = ?")
        params.append(client_id)
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    rows = conn.execute(
        f"""SELECT date, channel,
                  total_orders, failed_orders,
                  total_price_checks, inaccurate_checks
           FROM api_daily_metrics
           {where}
           ORDER BY date, channel""",
        params,
    ).fetchall()

    # avg_order_value from agoda_daily_metrics (same date/channel filters)
    agoda_cond: list = []
    agoda_params: list = []
    if start_date and end_date:
        agoda_cond.append("date BETWEEN ? AND ?")
        agoda_params.extend([start_date, end_date])
    if client_id:
        agoda_cond.append("client_id = ?")
        agoda_params.append(client_id)
    agoda_where = ("WHERE " + " AND ".join(agoda_cond)) if agoda_cond else ""
    avg_row = conn.execute(
        f"SELECT ROUND(SUM(ttv) / NULLIF(SUM(bookings), 0), 2) AS aov FROM agoda_daily_metrics {agoda_where}",
        agoda_params,
    ).fetchone()
    avg_order_value = avg_row["aov"] or 0

    conn.close()

    active_channels = [client_id] if client_id else _CHANNELS
    date_data: dict = defaultdict(dict)
    all_dates: list = []
    for row in rows:
        d = row["date"]
        if d not in all_dates:
            all_dates.append(d)
        date_data[d][row["channel"]] = {
            "to": row["total_orders"],
            "fo": row["failed_orders"],
            "tc": row["total_price_checks"],
            "ic": row["inaccurate_checks"],
        }

    def fmt_date(d: str) -> str:
        dt = datetime.strptime(d, "%Y-%m-%d")
        return f"{dt.strftime('%b')} {dt.day}"

    display_dates = [fmt_date(d) for d in all_dates]

    # per-channel accuracy trend
    accuracy_by_channel: dict = {ch: [] for ch in active_channels}
    pre_error_trend: list = []
    book_error_trend: list = []

    for d in all_dates:
        day = date_data[d]
        day_tc = sum(day.get(ch, {}).get("tc", 0) for ch in active_channels)
        day_ic = sum(day.get(ch, {}).get("ic", 0) for ch in active_channels)
        day_to = sum(day.get(ch, {}).get("to", 0) for ch in active_channels)
        day_fo = sum(day.get(ch, {}).get("fo", 0) for ch in active_channels)

        pre_error_trend.append(round(day_ic / day_tc * 100, 2) if day_tc else 0)
        book_error_trend.append(round(day_fo / day_to * 100, 2) if day_to else 0)

        for ch in active_channels:
            ch_d = day.get(ch, {})
            tc, ic = ch_d.get("tc", 0), ch_d.get("ic", 0)
            accuracy_by_channel[ch].append(round((tc - ic) / tc * 100, 1) if tc else 0)

    # per-channel book_error trend
    book_error_by_channel: dict = {ch: [] for ch in active_channels}
    for d in all_dates:
        day = date_data[d]
        for ch in active_channels:
            ch_d = day.get(ch, {})
            to_, fo = ch_d.get("to", 0), ch_d.get("fo", 0)
            book_error_by_channel[ch].append(round(fo / to_ * 100, 2) if to_ else 0)

    # overall summary
    total_tc = sum(r["total_price_checks"] for r in rows)
    total_ic = sum(r["inaccurate_checks"] for r in rows)
    total_to = sum(r["total_orders"] for r in rows)
    total_fo = sum(r["failed_orders"] for r in rows)

    return {
        "summary": {
            "pre_error_rate": round(total_ic / total_tc * 100, 2) if total_tc else 0,
            "book_error_rate": round(total_fo / total_to * 100, 2) if total_to else 0,
            "total_price_checks": total_tc,
            "total_orders": total_to,
            "estimated_ttv_loss": round(total_fo * avg_order_value),
        },
        "dates": display_dates,
        "accuracy_by_channel": accuracy_by_channel,
        "book_error_by_channel": book_error_by_channel,
        "pre_error_trend": pre_error_trend,
        "book_error_trend": book_error_trend,
    }


@app.post("/api/chat/dida-api")
def chat_dida_api(body: ChatRequest):
    import subprocess, shutil

    # Build conversation context for the prompt
    history = ""
    for m in body.messages[:-1]:
        role = "用户" if m.role == "user" else "助手"
        history += f"{role}: {m.content}\n\n"

    last_msg = body.messages[-1].content if body.messages else ""
    full_prompt = f"{DIDA_API_SYSTEM_PROMPT}\n\n{history}用户: {last_msg}\n\n助手:"

    claude_bin = shutil.which("claude") or "/home/jason.huang/.config/nvm/versions/node/v22.22.3/bin/claude"

    def generate():
        proc = subprocess.Popen(
            [claude_bin, "--print", "--model", "claude-sonnet-4-6"],
            stdin=subprocess.PIPE,
            stdout=subprocess.PIPE,
            stderr=subprocess.DEVNULL,
            text=True,
        )
        stdout, _ = proc.communicate(input=full_prompt, timeout=60)
        yield stdout

    return StreamingResponse(generate(), media_type="text/plain; charset=utf-8")


# ── Dida 联系方式（只读） ────────────────────────────────────────

@app.get("/api/contacts/dida")
def get_dida_contacts():
    conn = get_connection()
    contacts = conn.execute(
        "SELECT * FROM dida_contacts ORDER BY sort_order"
    ).fetchall()
    result = []
    for c in contacts:
        fields = conn.execute(
            "SELECT label, value FROM dida_contact_fields WHERE contact_id=? ORDER BY sort_order",
            (c["id"],),
        ).fetchall()
        result.append({
            "id": c["id"],
            "title": c["title"],
            "subtitle": c["subtitle"],
            "icon_key": c["icon_key"],
            "color": c["color"],
            "bg_color": c["bg_color"],
            "fields": [{"label": f["label"], "value": f["value"]} for f in fields],
        })
    conn.close()
    return result


# ── 我方对接人（CRUD） ───────────────────────────────────────────

class MyContactBody(BaseModel):
    type: str
    name: Optional[str] = ""
    role: Optional[str] = ""
    email: Optional[str] = ""
    phone: Optional[str] = ""
    wechat: Optional[str] = ""


@app.get("/api/contacts/my")
def get_my_contacts():
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM my_contacts ORDER BY type, sort_order, id"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/api/contacts/my", status_code=201)
def create_my_contact(body: MyContactBody):
    if body.type not in ("ops", "biz"):
        raise HTTPException(status_code=400, detail="type must be ops or biz")
    conn = get_connection()
    max_order = conn.execute(
        "SELECT COALESCE(MAX(sort_order),0) FROM my_contacts WHERE type=?", (body.type,)
    ).fetchone()[0]
    cur = conn.execute(
        "INSERT INTO my_contacts (type, name, role, email, phone, wechat, sort_order) VALUES (?,?,?,?,?,?,?)",
        (body.type, body.name, body.role, body.email, body.phone, body.wechat, max_order + 1),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM my_contacts WHERE id=?", (cur.lastrowid,)).fetchone()
    conn.close()
    return dict(row)


@app.put("/api/contacts/my/{contact_id}")
def update_my_contact(contact_id: int, body: MyContactBody):
    conn = get_connection()
    existing = conn.execute("SELECT id FROM my_contacts WHERE id=?", (contact_id,)).fetchone()
    if not existing:
        conn.close()
        raise HTTPException(status_code=404, detail="Contact not found")
    conn.execute(
        "UPDATE my_contacts SET type=?, name=?, role=?, email=?, phone=?, wechat=? WHERE id=?",
        (body.type, body.name, body.role, body.email, body.phone, body.wechat, contact_id),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM my_contacts WHERE id=?", (contact_id,)).fetchone()
    conn.close()
    return dict(row)


@app.delete("/api/contacts/my/{contact_id}", status_code=204)
def delete_my_contact(contact_id: int):
    conn = get_connection()
    conn.execute("DELETE FROM my_contacts WHERE id=?", (contact_id,))
    conn.commit()
    conn.close()



# ── 渠道匹配 ─────────────────────────────────────────────────────

@app.get("/api/channel-mapping")
def get_channel_mapping(
    dida_hotel_id: str = "",
    client_id: str = "",
    client_hotel_id: str = "",
):
    conn = get_connection()
    query = "SELECT * FROM channel_mappings WHERE 1=1"
    params: list = []
    if dida_hotel_id.strip():
        query += " AND CAST(dida_hotel_id AS TEXT) LIKE ?"
        params.append(f"%{dida_hotel_id.strip()}%")
    if client_id.strip():
        query += " AND client_id = ?"
        params.append(client_id.strip())
    if client_hotel_id.strip():
        query += " AND client_hotel_id LIKE ?"
        params.append(f"%{client_hotel_id.strip()}%")
    query += " ORDER BY id"
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/api/channel-mapping/upload")
async def upload_channel_mapping(file: UploadFile = File(...)):
    import openpyxl, io

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传 Excel 文件（.xlsx 或 .xls）")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Excel 文件解析失败，请检查文件格式")

    ws = wb.active
    upload_rows: list[tuple[int, str, str]] = []
    parse_errors: list[str] = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(c for c in row[:3] if c is not None):
            continue
        try:
            dida_id         = int(row[0])
            client_id_val   = str(row[1]).strip()
            client_hotel_val = str(row[2]).strip()
            if not client_id_val or not client_hotel_val:
                raise ValueError
            upload_rows.append((dida_id, client_id_val, client_hotel_val))
        except (TypeError, ValueError, IndexError):
            parse_errors.append(f"第 {i} 行数据格式错误（需要：DidaHotelID / 客户ID / 客户HotelID）")

    if parse_errors:
        raise HTTPException(status_code=400, detail="；".join(parse_errors[:5]))

    # 校验上传文件内部是否满足一一对应
    seen_dida_client: dict[tuple, str] = {}
    seen_client_hotel: dict[tuple, int] = {}
    internal_errors: list[str] = []
    for dida_id, cid, chid in upload_rows:
        k1 = (dida_id, cid)
        if k1 in seen_dida_client and seen_dida_client[k1] != chid:
            internal_errors.append(f"Dida Hotel ID {dida_id}（{cid}）在文件中存在多个不同映射")
        seen_dida_client[k1] = chid
        k2 = (cid, chid)
        if k2 in seen_client_hotel and seen_client_hotel[k2] != dida_id:
            internal_errors.append(f"客户 Hotel ID {chid}（{cid}）在文件中映射了多个不同 Dida Hotel ID")
        seen_client_hotel[k2] = dida_id

    if internal_errors:
        raise HTTPException(status_code=400, detail="文件中存在非一一对应关系：" + "；".join(internal_errors[:5]))

    # 逐行写入数据库
    conn = get_connection()
    added = 0
    conflict_errors: list[str] = []

    for dida_id, cid, chid in upload_rows:
        # 已存在完全相同的记录 → 跳过
        existing = conn.execute(
            "SELECT client_hotel_id FROM channel_mappings WHERE dida_hotel_id=? AND client_id=?",
            (dida_id, cid),
        ).fetchone()
        if existing:
            if existing["client_hotel_id"] == chid:
                continue  # 完全一致，跳过
            else:
                conflict_errors.append(
                    f"Dida Hotel ID {dida_id}（{cid}）已匹配 {existing['client_hotel_id']}，与上传的 {chid} 冲突"
                )
                continue

        # 反向：客户 Hotel ID 是否已映射到别的 Dida Hotel ID
        reverse = conn.execute(
            "SELECT dida_hotel_id FROM channel_mappings WHERE client_id=? AND client_hotel_id=?",
            (cid, chid),
        ).fetchone()
        if reverse:
            conflict_errors.append(
                f"客户 Hotel ID {chid}（{cid}）已映射到 Dida Hotel ID {reverse['dida_hotel_id']}，与上传的 {dida_id} 冲突"
            )
            continue

        conn.execute(
            "INSERT INTO channel_mappings (dida_hotel_id, client_id, client_hotel_id, updated_at) VALUES (?,?,?,datetime('now'))",
            (dida_id, cid, chid),
        )
        added += 1

    conn.commit()
    conn.close()
    return {"added": added, "conflicts": conflict_errors}


# ── 渠道热销 ─────────────────────────────────────────────────────

@app.get("/api/hot-sales")
def get_hot_sales(
    channel_id: str = "",
    hotel_id:   str = "",
    country:    str = "",
    city:       str = "",
):
    conn = get_connection()
    query = "SELECT * FROM channel_hot_sales WHERE 1=1"
    params: list = []
    if channel_id.strip():
        query += " AND channel_id = ?"
        params.append(channel_id.strip())
    if hotel_id.strip():
        query += " AND hotel_id LIKE ?"
        params.append(f"%{hotel_id.strip()}%")
    if country.strip():
        query += " AND country = ?"
        params.append(country.strip())
    if city.strip():
        query += " AND city LIKE ?"
        params.append(f"%{city.strip()}%")
    query += " ORDER BY id"
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.get("/api/hot-sales/stats")
def get_hot_sales_stats():
    conn = get_connection()
    total     = conn.execute("SELECT COUNT(*) FROM channel_hot_sales").fetchone()[0]
    countries = conn.execute("SELECT COUNT(DISTINCT country) FROM channel_hot_sales").fetchone()[0]
    conn.close()
    matched   = round(total * 0.54)
    return {"total": total, "matched": matched, "countries": countries}


@app.post("/api/hot-sales/upload")
async def upload_hot_sales(file: UploadFile = File(...)):
    import openpyxl, io

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传 Excel 文件（.xlsx 或 .xls）")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Excel 文件解析失败，请检查文件格式")

    ws = wb.active
    upload_rows: list[tuple] = []
    parse_errors: list[str] = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(c for c in row[:5] if c is not None):
            continue
        try:
            channel_id = str(row[0]).strip()
            hotel_id   = str(row[1]).strip()
            country    = str(row[2]).strip()
            city       = str(row[3]).strip()
            address    = str(row[4]).strip()
            if not channel_id or not hotel_id or not country or not city:
                raise ValueError
            upload_rows.append((channel_id, hotel_id, country, city, address))
        except (TypeError, ValueError, IndexError):
            parse_errors.append(f"第 {i} 行数据格式错误（需要：渠道ID / 酒店ID / 酒店国家 / 酒店城市 / 酒店地址）")

    if parse_errors:
        raise HTTPException(status_code=400, detail="；".join(parse_errors[:5]))

    # 校验文件内部无重复 (channel_id, hotel_id)
    seen: dict[tuple, str] = {}
    dup_errors: list[str] = []
    for channel_id, hotel_id, *_ in upload_rows:
        key = (channel_id, hotel_id)
        if key in seen:
            dup_errors.append(f"渠道 {channel_id} 的酒店 {hotel_id} 在文件中重复")
        seen[key] = hotel_id
    if dup_errors:
        raise HTTPException(status_code=400, detail="文件中存在重复数据：" + "；".join(dup_errors[:5]))

    conn = get_connection()
    added = 0
    skipped = 0

    for channel_id, hotel_id, country, city, address in upload_rows:
        existing = conn.execute(
            "SELECT id FROM channel_hot_sales WHERE channel_id=? AND hotel_id=?",
            (channel_id, hotel_id),
        ).fetchone()
        if existing:
            skipped += 1
            continue
        conn.execute(
            "INSERT INTO channel_hot_sales (channel_id, hotel_id, country, city, address, updated_at) VALUES (?,?,?,?,?,date('now'))",
            (channel_id, hotel_id, country, city, address),
        )
        added += 1

    conn.commit()
    conn.close()
    return {"added": added, "skipped": skipped}


# ── 渠道参数配置 ─────────────────────────────────────────────────────

@app.get("/api/channel-config")
def get_channel_config(client_id: Optional[str] = None):
    conn = get_connection()
    if client_id:
        rows = conn.execute(
            "SELECT * FROM channel_configurations WHERE client_id = ? ORDER BY id",
            (client_id,),
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM channel_configurations ORDER BY id"
        ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ── 订单日志 ──────────────────────────────────────────────────────────

@app.get("/api/order-logs")
def get_order_logs(
    order_no:     Optional[str] = None,
    order_status: Optional[str] = None,
    client_id:    Optional[str] = None,
):
    conn = get_connection()
    conditions = []
    params: list = []
    if order_no and order_no.strip():
        conditions.append("order_no LIKE ?")
        params.append(f"%{order_no.strip()}%")
    if order_status and order_status.strip():
        conditions.append("order_status = ?")
        params.append(order_status.strip())
    if client_id and client_id.strip():
        conditions.append("client_id = ?")
        params.append(client_id.strip())
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    rows = conn.execute(
        f"""SELECT order_no, client_id, order_status,
                  GROUP_CONCAT(log_type, '|') AS log_types,
                  MAX(updated_at) AS updated_at
           FROM order_logs
           {where}
           GROUP BY order_no, client_id, order_status
           ORDER BY MAX(updated_at) DESC
           LIMIT 50""",
        params,
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.get("/api/order-logs/{order_no}/detail")
def get_order_log_detail(order_no: str):
    import json as _json
    conn = get_connection()
    rows = conn.execute(
        """SELECT log_type, log_detail, updated_at
           FROM order_logs WHERE order_no = ? ORDER BY id""",
        (order_no,),
    ).fetchall()
    conn.close()
    if not rows:
        raise HTTPException(status_code=404, detail="订单不存在")
    return [{"log_type": r["log_type"],
             "log_detail": _json.loads(r["log_detail"]),
             "updated_at": r["updated_at"]} for r in rows]


# ── Admin: Account Management ────────────────────────────────────────────────

def _get_admin(authorization: Optional[str]) -> dict:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing token")
    token = authorization.split(" ", 1)[1]
    try:
        payload = verify_token(token)
    except ValueError:
        raise HTTPException(status_code=401, detail="Invalid token")
    conn = get_connection()
    row = conn.execute("SELECT * FROM users WHERE email = ?", (payload["email"],)).fetchone()
    conn.close()
    if not row or row["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return {"email": payload["email"], "id": int(payload["sub"])}


@app.get("/api/admin/users")
def admin_list_users(authorization: Optional[str] = Header(default=None)):
    admin = _get_admin(authorization)
    conn = get_connection()
    rows = conn.execute(
        "SELECT id, email, channel_name, contact_name, status, created_at"
        " FROM users WHERE created_by = ? ORDER BY id",
        (admin["email"],),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/api/admin/users")
def admin_create_user(body: CreateUserRequest, authorization: Optional[str] = Header(default=None)):
    admin = _get_admin(authorization)
    conn = get_connection()
    try:
        conn.execute(
            "INSERT INTO users (email, password_hash, channel_name, contact_name, role, status, created_by)"
            " VALUES (?,?,?,?,?,?,?)",
            (body.email, hash_password(body.password), body.channel_name, body.contact_name,
             "user", "active", admin["email"]),
        )
        conn.commit()
    except Exception:
        conn.close()
        raise HTTPException(status_code=400, detail="邮箱已存在")
    row = conn.execute(
        "SELECT id, email, channel_name, contact_name, status FROM users WHERE email = ?",
        (body.email,),
    ).fetchone()
    conn.close()
    return {"success": True, "user": dict(row)}


@app.patch("/api/admin/users/{user_id}/status")
def admin_toggle_status(user_id: int, authorization: Optional[str] = Header(default=None)):
    admin = _get_admin(authorization)
    conn = get_connection()
    row = conn.execute(
        "SELECT * FROM users WHERE id = ? AND created_by = ?", (user_id, admin["email"])
    ).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="用户不存在")
    new_status = "disabled" if row["status"] == "active" else "active"
    conn.execute("UPDATE users SET status = ? WHERE id = ?", (new_status, user_id))
    conn.commit()
    conn.close()
    return {"success": True, "status": new_status}


@app.delete("/api/admin/users/{user_id}")
def admin_delete_user(user_id: int, authorization: Optional[str] = Header(default=None)):
    admin = _get_admin(authorization)
    conn = get_connection()
    row = conn.execute(
        "SELECT * FROM users WHERE id = ? AND created_by = ?", (user_id, admin["email"])
    ).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="用户不存在")
    conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()
    return {"success": True}


# ── 转化指标 ─────────────────────────────────────────────────────

_CONV_CHANNELS = ["Agoda", "AgodaEBK", "AgodaUK", "Lvzan", "DidaOpaq", "Barli2b"]


@app.get("/api/conversion/metrics")
def conversion_metrics(
    start_date: Optional[str] = Query(default=None),
    end_date:   Optional[str] = Query(default=None),
    client_id:  Optional[str] = Query(default=None),
):
    from collections import defaultdict
    from datetime import datetime

    active_channels = [client_id] if client_id else _CONV_CHANNELS

    cond: list = []
    params: list = []
    if start_date:
        cond.append("date >= ?")
        params.append(start_date)
    if end_date:
        cond.append("date <= ?")
        params.append(end_date)
    if client_id:
        cond.append("channel = ?")
        params.append(client_id)
    where = ("WHERE " + " AND ".join(cond)) if cond else ""

    conn = get_connection()
    rows = conn.execute(
        f"""SELECT date, channel, look, property, avail_look, prebook, book
           FROM conversion_daily_metrics
           {where}
           ORDER BY date, channel""",
        params,
    ).fetchall()
    conn.close()

    date_data: dict = defaultdict(dict)
    all_dates: list = []
    for r in rows:
        d = r["date"]
        if d not in all_dates:
            all_dates.append(d)
        date_data[d][r["channel"]] = dict(r)

    def fmt(d: str) -> str:
        dt = datetime.strptime(d, "%Y-%m-%d")
        return f"{dt.strftime('%b')} {dt.day}"

    display_dates = [fmt(d) for d in all_dates]

    # 每渠道每日趋势
    trends: dict = {ch: {"l2b": [], "p2b": [], "l2c": [], "c2b": []} for ch in active_channels}
    for d in all_dates:
        day = date_data[d]
        for ch in active_channels:
            v = day.get(ch, {})
            look  = v.get("look",       0)
            prop  = v.get("property",   0)
            avail = v.get("avail_look", 0)
            pre   = v.get("prebook",    0)
            book  = v.get("book",       0)
            trends[ch]["l2b"].append(round(look  / book, 1) if book else None)
            trends[ch]["p2b"].append(round(prop  / book, 1) if book else None)
            trends[ch]["l2c"].append(round(avail / pre,  1) if pre  else None)
            trends[ch]["c2b"].append(round(pre   / book, 1) if book else None)

    # 全渠道汇总（sum/sum）
    totals = {"look": 0, "property": 0, "avail_look": 0, "prebook": 0, "book": 0}
    for r in rows:
        for k in totals:
            totals[k] += r[k]

    def safe(a, b):
        return round(a / b, 1) if b else None

    summary = {
        "l2b": safe(totals["look"],       totals["book"]),
        "p2b": safe(totals["property"],   totals["book"]),
        "l2c": safe(totals["avail_look"], totals["prebook"]),
        "c2b": safe(totals["prebook"],    totals["book"]),
    }

    # 全渠道每日汇总趋势
    agg_trend: dict = {"l2b": [], "p2b": [], "l2c": [], "c2b": []}
    for d in all_dates:
        day = date_data[d]
        lk = sum(day.get(ch, {}).get("look",       0) for ch in active_channels)
        pp = sum(day.get(ch, {}).get("property",   0) for ch in active_channels)
        av = sum(day.get(ch, {}).get("avail_look", 0) for ch in active_channels)
        pb = sum(day.get(ch, {}).get("prebook",    0) for ch in active_channels)
        bk = sum(day.get(ch, {}).get("book",       0) for ch in active_channels)
        agg_trend["l2b"].append(round(lk / bk, 1) if bk else None)
        agg_trend["p2b"].append(round(pp / bk, 1) if bk else None)
        agg_trend["l2c"].append(round(av / pb, 1) if pb else None)
        agg_trend["c2b"].append(round(pb / bk, 1) if bk else None)

    return {
        "dates":     display_dates,
        "summary":   summary,
        "trends":    trends,
        "agg_trend": agg_trend,
    }


# ── 错误日志 ──────────────────────────────────────────────────────

@app.get("/api/errors/prebook")
def errors_prebook(
    client_id:    str = Query(""),
    error_type:   str = Query(""),
    rate_plan_id: str = Query(""),
    page:         int = Query(1, ge=1),
    page_size:    int = Query(15, ge=1, le=100),
):
    conn = get_connection()

    # 动态筛选条件
    where, params = ["1=1"], []
    if client_id:
        where.append("client_id = ?");    params.append(client_id)
    if error_type:
        where.append("error_type = ?");   params.append(error_type)
    if rate_plan_id:
        where.append("dida_rate_plan_id LIKE ?"); params.append(f"%{rate_plan_id}%")

    w = " AND ".join(where)

    # 图表数据（筛选后按错误类型汇总）
    chart_rows = conn.execute(
        f"SELECT error_type, COUNT(*) AS cnt FROM prebook_error_logs WHERE {w} GROUP BY error_type ORDER BY cnt DESC",
        params,
    ).fetchall()
    chart = [{"error_type": r["error_type"], "count": r["cnt"]} for r in chart_rows]

    # 总数
    total = conn.execute(f"SELECT COUNT(*) FROM prebook_error_logs WHERE {w}", params).fetchone()[0]

    # 分页明细
    offset = (page - 1) * page_size
    rows = conn.execute(
        f"SELECT log_time, client_id, dida_rate_plan_id, dida_hotel_id, error_type, rate_record_channel"
        f" FROM prebook_error_logs WHERE {w} ORDER BY log_time DESC LIMIT ? OFFSET ?",
        params + [page_size, offset],
    ).fetchall()

    conn.close()
    return {
        "chart": chart,
        "total": total,
        "page": page,
        "page_size": page_size,
        "rows": [dict(r) for r in rows],
    }


@app.get("/api/errors/book")
def errors_book(
    client_id:      str = Query(""),
    error_type:     str = Query(""),
    booking_number: str = Query(""),
    page:           int = Query(1, ge=1),
    page_size:      int = Query(15, ge=1, le=100),
):
    conn = get_connection()

    where, params = ["1=1"], []
    if client_id:
        where.append("client_id = ?");        params.append(client_id)
    if error_type:
        where.append("error_type = ?");       params.append(error_type)
    if booking_number:
        where.append("channel_bookingnumber LIKE ?"); params.append(f"%{booking_number}%")

    w = " AND ".join(where)

    chart_rows = conn.execute(
        f"SELECT error_type, COUNT(*) AS cnt FROM book_error_logs WHERE {w} GROUP BY error_type ORDER BY cnt DESC",
        params,
    ).fetchall()
    chart = [{"error_type": r["error_type"], "count": r["cnt"]} for r in chart_rows]

    total = conn.execute(f"SELECT COUNT(*) FROM book_error_logs WHERE {w}", params).fetchone()[0]

    offset = (page - 1) * page_size
    rows = conn.execute(
        f"SELECT channel_createtime, client_id, channel_bookingnumber, dida_hotel_id, error_type"
        f" FROM book_error_logs WHERE {w} ORDER BY channel_createtime DESC LIMIT ? OFFSET ?",
        params + [page_size, offset],
    ).fetchall()

    conn.close()
    return {
        "chart": chart,
        "total": total,
        "page": page,
        "page_size": page_size,
        "rows": [dict(r) for r in rows],
    }


@app.get("/api/errors/meta")
def errors_meta():
    """返回下拉框所需的渠道和错误类型列表"""
    conn = get_connection()
    pre_clients  = [r[0] for r in conn.execute("SELECT DISTINCT client_id FROM prebook_error_logs ORDER BY client_id").fetchall()]
    pre_errors   = [r[0] for r in conn.execute("SELECT DISTINCT error_type  FROM prebook_error_logs ORDER BY error_type").fetchall()]
    book_clients = [r[0] for r in conn.execute("SELECT DISTINCT client_id FROM book_error_logs ORDER BY client_id").fetchall()]
    book_errors  = [r[0] for r in conn.execute("SELECT DISTINCT error_type  FROM book_error_logs ORDER BY error_type").fetchall()]
    conn.close()
    return {
        "prebook": {"channels": pre_clients,  "error_types": pre_errors},
        "book":    {"channels": book_clients, "error_types": book_errors},
    }
