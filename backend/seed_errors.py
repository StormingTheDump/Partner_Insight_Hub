"""
从 Excel 示例文件采样生成错误日志 demo 数据。
- 验价：采样 900 条，修改 dida_rate_plan_id 末尾数字
- 下单：全部有效条，修改 channel_bookingnumber 末尾数字
重复运行安全。
"""
import sys, os, random
sys.path.insert(0, os.path.dirname(__file__))

from database import get_connection

PREBOOK_XLSX = os.path.expanduser("~/Downloads/报错验价底表示例.xlsx")
BOOK_XLSX    = os.path.expanduser("~/Downloads/报错下单底表示例.xlsx")

PREBOOK_SAMPLE = 900
RNG = random.Random(42)


def _modify_rate_plan_id(rpid: str) -> str:
    try:
        n = int(rpid)
        offset = RNG.randint(-499, 499)
        return str(n + offset)
    except (ValueError, TypeError):
        return str(rpid)


def _modify_booking_number(bn) -> str:
    try:
        n = int(str(bn))
        offset = RNG.randint(-99, 99)
        return str(max(10_000_000_000, n + offset))
    except (ValueError, TypeError):
        return str(bn)


def seed():
    try:
        import openpyxl
    except ImportError:
        print("请先安装 openpyxl: pip install openpyxl")
        return

    conn = get_connection()

    # ── 验价报错 ─────────────────────────────────────────────────
    pre_count = conn.execute("SELECT COUNT(*) FROM prebook_error_logs").fetchone()[0]
    if pre_count == 0:
        wb = openpyxl.load_workbook(PREBOOK_XLSX, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(min_row=2, values_only=True))
        sample = RNG.sample(all_rows, min(PREBOOK_SAMPLE, len(all_rows)))

        rows = []
        for r in sample:
            log_time, client_id, rate_plan_id, hotel_id, error_type, rate_record = r
            if not error_type:
                continue
            log_time_str = str(log_time).replace("+08", "").strip() if log_time else None
            rows.append((
                log_time_str,
                client_id,
                _modify_rate_plan_id(str(rate_plan_id)) if rate_plan_id else None,
                hotel_id,
                error_type,
                rate_record,
            ))

        conn.executemany(
            "INSERT INTO prebook_error_logs"
            " (log_time, client_id, dida_rate_plan_id, dida_hotel_id, error_type, rate_record_channel)"
            " VALUES (?,?,?,?,?,?)",
            rows,
        )
        conn.commit()
        print(f"prebook_error_logs: 写入 {len(rows)} 行")
    else:
        print(f"prebook_error_logs: 已有 {pre_count} 行，跳过")

    # ── 下单报错 ─────────────────────────────────────────────────
    book_count = conn.execute("SELECT COUNT(*) FROM book_error_logs").fetchone()[0]
    if book_count == 0:
        wb2 = openpyxl.load_workbook(BOOK_XLSX, data_only=True)
        ws2 = wb2.active
        all_rows2 = list(ws2.iter_rows(min_row=2, values_only=True))

        rows2 = []
        for r in all_rows2:
            createtime, client_id, booking_no, hotel_id, error_type = r
            if not error_type:
                continue
            ct_str = createtime.strftime("%Y-%m-%d %H:%M:%S") if hasattr(createtime, "strftime") else str(createtime)
            rows2.append((
                ct_str,
                client_id,
                _modify_booking_number(booking_no),
                hotel_id,
                error_type,
            ))

        conn.executemany(
            "INSERT INTO book_error_logs"
            " (channel_createtime, client_id, channel_bookingnumber, dida_hotel_id, error_type)"
            " VALUES (?,?,?,?,?)",
            rows2,
        )
        conn.commit()
        print(f"book_error_logs: 写入 {len(rows2)} 行")
    else:
        print(f"book_error_logs: 已有 {book_count} 行，跳过")

    conn.close()


if __name__ == "__main__":
    seed()
